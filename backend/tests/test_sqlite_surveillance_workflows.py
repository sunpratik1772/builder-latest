from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from data_sources.registry import get_registry
from data_sources.sqlite_demo import ALL_TABLES, DEFAULT_DB_PATH, alert_payload
from engine.context import RunContext
from engine.dag_runner import run_workflow
from engine.nodes.db_market_connector import handle_db_market_connector
from engine.nodes.db_solr_connector import handle_db_solr_connector
from engine.nodes.market_api_connector import handle_market_api_connector


WORKFLOW = Path(__file__).resolve().parents[1] / "workflows" / "sqlite_surveillance_e2e_workflow.json"


class _FakeAdapter:
    def single_shot(self, prompt: str, **_: object) -> str:
        verdict = "evidence is explainable or insufficient for front-running"
        if "front_running_positive" in prompt:
            verdict = "likely front-running based on aligned order, book, market, and comms evidence"
        return f"Gemini summary: {verdict}. Evidence: {prompt[:240]}"


def _workflow(alert_id: str = "ALERT-FR-001") -> dict:
    dag = json.loads(WORKFLOW.read_text())
    for node in dag["nodes"]:
        if node["id"] == "workflow_context":
            node["config"] = {"scenario": "fxfro", "alert_id": alert_id}
            break
    return dag


def _workflow_with_context(config: dict) -> dict:
    dag = json.loads(WORKFLOW.read_text())
    for node in dag["nodes"]:
        if node["id"] == "workflow_context":
            node["config"] = config
            break
    return dag


def _safe_sheet(value: str) -> str:
    return re.sub(r"[\\/*?:\[\]]+", "-", str(value))[:18]


def _order_sheet_name(order_id: str) -> str:
    return f"order summary_{_safe_sheet(order_id)}"


def _book_sheet_name(book: str) -> str:
    return f"{_safe_sheet(book)} summary"


def _market_sheet_name(order_id: str) -> str:
    return f"market_datafor{order_id}"[:31]


def _comms_sheet_name(order_id: str) -> str:
    return f"commsfor {order_id}"[:31]


def test_sqlite_fixture_has_exact_six_tables_with_rows():
    assert DEFAULT_DB_PATH.exists()
    with sqlite3.connect(DEFAULT_DB_PATH) as conn:
        tables = {
            row[0]
            for row in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        }
        assert tables == set(ALL_TABLES)
        for table in ALL_TABLES:
            count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            assert count == 500, table
        positive = conn.execute(
            "SELECT COUNT(*) FROM hs_alerts WHERE scenario = 'front_running_positive'"
        ).fetchone()[0]
        negative = conn.execute(
            "SELECT COUNT(*) FROM hs_alerts WHERE scenario = 'front_running_negative'"
        ).fetchone()[0]
        assert positive > 3
        assert negative > 2


def test_registry_loads_exact_sqlite_datasources():
    ids = {source.id for source in get_registry().all()}
    assert ids == set(ALL_TABLES)
    assert "book" in get_registry().get("hs_orders").column_names()
    assert "display_post" in get_registry().get("comms_messages").column_names()


def test_demo_data_api_scopes_by_participant_keyword_and_date():
    client = TestClient(app)
    response = client.get(
        "/api/demo-data/search",
        params={
            "participant_id": "P-T002",
            "keyword": "client flow",
            "date": "2024-01-02",
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()["datasets"]
    assert {row["alert_id"] for row in payload["hs_alerts"]} == {"ALERT-FR-002"}
    assert {row["participant_id"] for row in payload["hs_trades"]} == {"P-T002"}
    assert {row["currency_pair"] for row in payload["market_ticks"]} == {"GBP/USD"}
    assert all("client flow" in row["display_post"] for row in payload["comms_messages"])


def test_workflow_nodes_are_registered_in_integration_palette():
    client = TestClient(app)
    manifest = client.get("/node-manifest").json()
    by_type = {node["type_id"]: node for node in manifest["nodes"]}
    assert by_type["WORKFLOW_CONTEXT"]["display_name"] == "workflow_context"
    assert by_type["FAN_OUT"]["palette_group"] == "flow"
    for type_id, display_name in {
        "DB_SOLR_CONNECTOR": "db_solr_connector",
        "DB_MARKET_CONNECTOR": "db_market_connector",
        "MARKET_API_CONNECTOR": "market_api_connector",
    }.items():
        node = by_type[type_id]
        assert node["palette_group"] == "int"
        assert node["display_name"] == display_name


def test_sqlite_connectors_filter_selected_output_columns():
    ctx = RunContext()
    ctx.alert_payload.update(alert_payload("ALERT-FR-001"))

    handle_db_solr_connector(
        {
            "id": "orders",
            "config": {
                "table": "hs_orders",
                "output_name": "orders",
                "output_columns": ["order_id", "book", "limit_price"],
            },
        },
        ctx,
    )
    assert set(ctx.get("orders_output")[0]) == {"order_id", "book", "limit_price", "_dataset"}
    assert list(ctx.datasets["orders"].columns) == ["order_id", "book", "limit_price"]

    handle_db_market_connector(
        {
            "id": "market",
            "config": {
                "output_name": "market",
                "output_columns": ["timestamp", "mid"],
                "limit": 2,
            },
        },
        ctx,
    )
    assert set(ctx.get("market_output")[0]) == {"timestamp", "mid", "_dataset"}
    assert len(ctx.get("market_output")) == 2

    handle_market_api_connector(
        {
            "id": "comms",
            "config": {
                "datasets": ["comms_messages"],
                "output_columns": ["timestamp", "display_post"],
                "limit": 1,
            },
        },
        ctx,
    )
    assert set(ctx.get("comms_output")[0]) == {"timestamp", "display_post", "_dataset"}
    assert len(ctx.get("comms_output")) == 1


def test_definition_of_done_for_five_alert_ids(monkeypatch):
    from engine.nodes import llm_basic

    monkeypatch.setattr(llm_basic, "get_default_adapter", lambda: _FakeAdapter())
    scenarios = {
        "ALERT-FR-001": {"books": 3, "positive": True},
        "ALERT-FR-002": {"books": 4, "positive": True},
        "ALERT-FR-003": {"books": 2, "positive": True},
        "ALERT-FR-004": {"books": 3, "positive": False},
        "ALERT-FR-005": {"books": 5, "positive": False},
    }
    summaries: dict[str, str] = {}
    metric_signatures: dict[str, tuple[object, ...]] = {}
    client = TestClient(app)
    base_expected_sheets = {"alert summary", "overall summary"}
    for alert_id, expected in scenarios.items():
        ctx = run_workflow(_workflow(alert_id), {})
        assert ctx.report_path
        path = Path(ctx.report_path)
        assert path.exists()
        assert path.suffix == ".xlsx"
        assert ctx.executive_summary
        assert ctx.sections
        assert ctx.get("book_count") == expected["books"]
        order_ids = ctx.get("order_ids", [])
        book_sheets = {_book_sheet_name(book) for book in ctx.get("distinct_books", [])}
        order_sheets = {_order_sheet_name(order_id) for order_id in order_ids}
        market_sheets = {_market_sheet_name(order_id) for order_id in order_ids}
        comms_sheets = {_comms_sheet_name(order_id) for order_id in order_ids}
        expected_sheets = base_expected_sheets | book_sheets | order_sheets | market_sheets | comms_sheets
        assert len(book_sheets) == expected["books"]
        assert len(order_sheets) == ctx.get("order_count")
        assert len(market_sheets) == ctx.get("order_count")
        assert len(comms_sheets) == ctx.get("order_count")

        download = client.get(f"/report/{path.name}")
        assert download.status_code == 200, download.text
        assert download.content[:4] == b"PK\x03\x04"

        wb = load_workbook(path)
        assert set(wb.sheetnames) == expected_sheets
        assert all(wb[name].max_row >= 2 for name in wb.sheetnames)

        for sheet in expected_sheets:
            if sheet == "overall summary":
                continue
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]
            alert_col = headers.index("alert_id") + 1
            row_alerts = {
                ws.cell(row=row_idx, column=alert_col).value
                for row_idx in range(2, ws.max_row + 1)
                if ws.cell(row=row_idx, column=alert_col).value
            }
            assert row_alerts == {alert_id}

        def sheet_value(sheet: str, column: str) -> object:
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]
            col_idx = headers.index(column) + 1
            return ws.cell(row=2, column=col_idx).value

        first_order = order_ids[0]
        first_book = ctx.get("distinct_books", [])[0]

        metric_signatures[alert_id] = (
            ctx.get("order_count"),
            ctx.get("book_count"),
            sheet_value(_order_sheet_name(first_order), "quantity"),
            sheet_value(_book_sheet_name(first_book), "max_notional_usd"),
            sheet_value(_market_sheet_name(first_order), "mid_move"),
            sheet_value(_comms_sheet_name(first_order), "keyword_hit_count"),
        )

        highlighted = 0
        for sheet in book_sheets | comms_sheets:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2):
                if any(cell.fill.fill_type == "solid" for cell in row):
                    highlighted += 1
        assert highlighted > 0

        final_ws = wb["overall summary"]
        values = [
            str(cell.value)
            for row in final_ws.iter_rows(min_row=2)
            for cell in row
            if cell.value is not None
        ]
        final_text = " ".join(values)
        summaries[alert_id] = final_text
        assert "Gemini summary" in final_text
        response = ctx.get("response_output", [])[0]["response"]
        assert response["artifact"] == str(path)
        assert response["alert_id"] == alert_id
        assert set(response["llm_summary"]) == expected_sheets
        assert all(response["llm_summary"].values())
        assert set(response["order_summary"]) == order_sheets
        assert response["overall_summary"]
        assert "overall summary" in response["llm_summary"]
        assert response["overall_summary"] == response["llm_summary"]["overall summary"]
        if expected["positive"]:
            assert "likely front-running" in final_text
        else:
            assert "explainable or insufficient" in final_text
        assert any("summary" in str(header).lower() for header in [cell.value for cell in final_ws[1]])

    assert len(set(summaries.values())) == len(scenarios)
    assert len(set(metric_signatures.values())) == len(scenarios)


# Exact workflow_context payloads (scenario + alert_id + hydrated fields) for regression.
_FULL_PAYLOAD_FIXTURES: list[tuple[str, dict]] = [
    (
        "ALERT-FR-001",
        {
            "scenario": "fxfro",
            "alert_id": "ALERT-FR-001",
            "participant_id": "P-T001",
            "trader_id": "T001",
            "trader_name": "Avery Shah",
            "keyword": "fixing",
            "currency_pair": "EUR/USD",
            "date": "2024-01-01",
            "alert_date": "2024-01-01",
            "start_time": "2024-01-01T11:07:00",
            "end_time": "2024-01-01T11:52:00",
        },
    ),
    (
        "ALERT-FR-002",
        {
            "scenario": "fxfro",
            "alert_id": "ALERT-FR-002",
            "participant_id": "P-T002",
            "trader_id": "T002",
            "trader_name": "Morgan Lee",
            "keyword": "client flow",
            "currency_pair": "GBP/USD",
            "date": "2024-01-02",
            "alert_date": "2024-01-02",
            "start_time": "2024-01-02T14:14:00",
            "end_time": "2024-01-02T14:59:00",
        },
    ),
    (
        "ALERT-FR-003",
        {
            "scenario": "fxfro",
            "alert_id": "ALERT-FR-003",
            "participant_id": "P-T003",
            "trader_id": "T003",
            "trader_name": "Riley Chen",
            "keyword": "large order",
            "currency_pair": "USD/JPY",
            "date": "2024-01-03",
            "alert_date": "2024-01-03",
            "start_time": "2024-01-03T09:21:00",
            "end_time": "2024-01-03T10:06:00",
        },
    ),
    (
        "ALERT-FR-004",
        {
            "scenario": "fxfro",
            "alert_id": "ALERT-FR-004",
            "participant_id": "P-T004",
            "trader_id": "T004",
            "trader_name": "Sam Patel",
            "keyword": "pre hedge",
            "currency_pair": "AUD/USD",
            "date": "2024-01-04",
            "alert_date": "2024-01-04",
            "start_time": "2024-01-04T12:28:00",
            "end_time": "2024-01-04T13:13:00",
        },
    ),
    (
        "ALERT-FR-005",
        {
            "scenario": "fxfro",
            "alert_id": "ALERT-FR-005",
            "participant_id": "P-T005",
            "trader_id": "T005",
            "trader_name": "Jordan Blake",
            "keyword": "risk transfer",
            "currency_pair": "USD/CHF",
            "date": "2024-01-05",
            "alert_date": "2024-01-05",
            "start_time": "2024-01-05T15:35:00",
            "end_time": "2024-01-05T16:20:00",
        },
    ),
]


def test_full_payload_contexts_match_user_fixtures(monkeypatch):
    """Run surveillance E2E with the exact fxfro payloads (all five alert ids)."""
    from engine.nodes import llm_basic

    monkeypatch.setattr(llm_basic, "get_default_adapter", lambda: _FakeAdapter())
    for alert_id, payload in _FULL_PAYLOAD_FIXTURES:
        ctx = run_workflow(_workflow_with_context(dict(payload)), {})
        response = ctx.get("response_output", [])[0]["response"]

        assert response["alert_id"] == alert_id
        assert ctx.alert_payload.get("alert_id") == alert_id
        assert ctx.alert_payload.get("currency_pair") == payload["currency_pair"]
        assert Path(response["artifact"]).exists()
        assert "overall summary" in response["llm_summary"]
        assert response["overall_summary"] == response["llm_summary"]["overall summary"]
        assert len([name for name in response["llm_summary"] if name.startswith("order summary_")]) == ctx.get(
            "order_count"
        )
        assert len(response["order_summary"]) == ctx.get("order_count")
        assert len([name for name in response["llm_summary"] if name.endswith(" summary")]) >= ctx.get("book_count")
