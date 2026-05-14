"""SPREADSHEET_FILE Node"""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from ..context import RunContext
from ..node_spec import NodeSpec, _spec_from_yaml


def _to_rows(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        rows = value
    elif value is None:
        rows = []
    else:
        rows = [value]
    out: list[dict[str, Any]] = []
    for row in rows:
        if isinstance(row, dict):
            if "json" in row and isinstance(row["json"], dict):
                out.append(row["json"])
            else:
                out.append(row)
        else:
            out.append({"value": row})
    return out


def _sheet_headers(rows: list[dict[str, Any]]) -> list[str]:
    return sorted({k for r in rows for k in r.keys()})


def _excel_cell(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _write_xlsx(
    *,
    file_path: Path,
    rows: list[dict[str, Any]],
    sheet_name: str,
    append: bool,
) -> int:
    if append and file_path.exists():
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        # Drop default blank sheet on first write.
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    title = (sheet_name or "Sheet1")[:31]
    if title in wb.sheetnames:
        ws = wb[title]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=title)

    headers = _sheet_headers(rows)
    if headers:
        ws.append(headers)
        for row in rows:
            ws.append([_excel_cell(row.get(h)) for h in headers])

    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    return file_path.stat().st_size


def _read_xlsx(file_path: Path) -> list[dict[str, Any]]:
    if not file_path.exists():
        return []
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if not wb.sheetnames:
        return []
    ws = wb[wb.sheetnames[0]]
    data = list(ws.iter_rows(values_only=True))
    if not data:
        return []
    headers = [str(h) if h is not None else "" for h in data[0]]
    out: list[dict[str, Any]] = []
    for row in data[1:]:
        out.append({headers[idx]: row[idx] for idx in range(len(headers))})
    return out


def handle_spreadsheet_file(node: dict, ctx: RunContext) -> None:
    """Read/write spreadsheet files."""
    cfg = node.get("config", {}) or {}
    node_id = node.get("id", "spreadsheet_file")

    operation = str(cfg.get("operation", "read")).lower()
    file_format = str(cfg.get("file_format", "csv")).lower()
    file_path_s = str(cfg.get("file_path", cfg.get("path", "")) or "")
    file_name = str(cfg.get("file_name", f"{node_id}.{file_format}"))
    if not file_path_s:
        file_path_s = f"/tmp/{file_name}"
    file_path = Path(file_path_s).expanduser()

    if operation == "read":
        if file_format == "xlsx" or file_path.suffix.lower() == ".xlsx":
            result = _read_xlsx(file_path)
        elif file_format == "json" or file_path.suffix.lower() == ".json":
            result = json.loads(file_path.read_text(encoding="utf-8")) if file_path.exists() else []
        else:
            if file_path.exists():
                text = file_path.read_text(encoding="utf-8", errors="ignore")
                result = list(csv.DictReader(text.splitlines())) if text.strip() else []
            else:
                result = []
        ctx.set(f"{node_id}_output", result)
        return

    input_items = _to_rows(ctx.get(f"{node_id}_input", []))
    options = cfg.get("options", {}) or {}
    sheet_name = str(options.get("sheet_name", options.get("sheetName", "Sheet1")))
    append = bool(options.get("append_to_existing_file", options.get("appendToExistingFile", False)))

    if file_format == "xlsx" or file_path.suffix.lower() == ".xlsx":
        bytes_written = _write_xlsx(
            file_path=file_path,
            rows=input_items,
            sheet_name=sheet_name,
            append=append,
        )
    elif file_format == "json" or file_path.suffix.lower() == ".json":
        file_path.parent.mkdir(parents=True, exist_ok=True)
        file_path.write_text(json.dumps(input_items, ensure_ascii=False, indent=2), encoding="utf-8")
        bytes_written = file_path.stat().st_size
    else:
        headers = _sheet_headers(input_items)
        file_path.parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            if headers:
                w = csv.DictWriter(f, fieldnames=headers)
                w.writeheader()
                for row in input_items:
                    w.writerow(row)
        bytes_written = file_path.stat().st_size

    result = [{
        "status": "written",
        "count": len(input_items),
        "path": str(file_path),
        "file_format": file_format,
        "bytes": bytes_written,
        "sheet_name": sheet_name if file_format == "xlsx" else None,
    }]
    ctx.set(f"{node_id}_output", result)


NODE_SPEC: NodeSpec = _spec_from_yaml(Path(__file__).with_suffix(".yaml"), handle_spreadsheet_file)